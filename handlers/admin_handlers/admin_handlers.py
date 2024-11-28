import os

import openpyxl
from aiogram import types  # Типы пользователя
from aiogram.filters import Command
from openpyxl.utils import get_column_letter
from aiogram.types import FSInputFile
from database.database import get_export_bonus_from_database
from database.database import get_export_user_bonus_from_database
from system.dispatcher import dp, bot  # Подключение к боту и диспетчеру пользователя
from system.dispatcher import router


@router.message(Command("export_bonus"))
async def export_command(message: types.Message):
    """Обработчик команды /export_bonus"""

    # Проверяем, является ли пользователь, который вызывает команду, администратором
    if message.from_user.id not in [5837917794, 5958542955,
                                    535185511]:  # Предоставление доступа к команде  /export_bonus
        await message.reply('У вас нет доступа к этой команде.', parse_mode='HTML')
        return

    data = get_export_bonus_from_database()  # Получаем данные бонусов

    wb = openpyxl.Workbook()  # Создаем файл Excel и записываем данные
    sheet = wb.active
    # Записываем заголовки
    headers = ['user_key', 'id', 'full_name', 'user_name', 'bonus', 'plase']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header
    # Записываем данные пользователей
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            sheet[f'{col_letter}{row_num}'] = cell_data

    wb.save('users_bonus.xlsx')  # Сохраняем файл Excel

    file = FSInputFile(f'users_bonus.xlsx')  # Отправляем файл пользователю

    await bot.send_document(message.from_user.id, document=file, caption='Данные пользователей в формате Excel',
                            parse_mode="HTML")  # Отправка файла пользователю

    os.remove('users_bonus.xlsx')  # Удаляем файл Excel


@router.message(Command("export_user"))
async def export_command(message: types.Message):
    """Обработчик команды /export_user"""
    # Проверяем, является ли пользователь, который вызывает команду, администратором
    if message.from_user.id not in [5837917794, 5958542955,
                                    535185511]:  # Предоставление доступа к команде  /export_user
        await message.reply('У вас нет доступа к этой команде.')
        return

    data = get_export_user_bonus_from_database()

    wb = openpyxl.Workbook()  # Создаем файл Excel и записываем данные
    sheet = wb.active
    # Записываем заголовки
    headers = ['ID', 'User ID', 'First Name', 'Last Name', 'Username', 'Date']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header
    # Записываем данные пользователей
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            sheet[f'{col_letter}{row_num}'] = cell_data

    wb.save('users.xlsx')  # Сохраняем файл Excel

    file = FSInputFile(f'users.xlsx')  # Отправляем файл пользователю
    await bot.send_document(message.from_user.id, document=file, caption='Данные пользователей в формате Excel',
                            parse_mode="HTML")  # Отправка файла пользователю

    os.remove('users.xlsx')  # Удаляем файл Excel


def register_admin_handler():
    """Регистрируем handlers для админа"""
    dp.register_message_handler(export_command)
