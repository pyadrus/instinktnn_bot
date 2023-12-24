import datetime  # Дата
import os
import random
import sqlite3  # Импортируем модуль для работы с базой данных SQLite
import time
from datetime import date

import openpyxl
from aiogram import types  # Типы пользователя
from aiogram.dispatcher import FSMContext  # Состояния пользователя
from aiogram.dispatcher.filters.state import StatesGroup, State
from loguru import logger
from openpyxl.utils import get_column_letter

from database.database import get_export_bonus_from_database
from database.database import get_export_user_bonus_from_database
from database.database import recording_the_data_of_users_who_launched_the_bot
from database.database import retrieve_user_bonus
from keyboards.bonus_keyboards import bonus_keyboards, top_kub_keyboards, bottom_kub_keyboards
from keyboards.greeting_keyboards import city_selection_keyboard  # Клавиатуры поста приветствия
from messages.bonus_text import random_bon, bonus_post
from system.dispatcher import dp, bot  # Подключение к боту и диспетчеру пользователя

logger.add('log/log.log', rotation='2 MB')


@dp.message_handler(commands=['start'])
async def greeting(message: types.Message, state: FSMContext):
    """Обработчик команды /start, он же пост приветствия"""
    await state.finish()
    await state.reset_state()
    current_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # Получаем текущую дату и время
    recording_the_data_of_users_who_launched_the_bot(message, current_date)
    logger.info(f'Привет! нажали на кнопку /start {message.from_user.id, message.from_user.username, current_date}')
    city_selection_key = city_selection_keyboard()
    await message.answer('Привет! 👋\n\n 🌟 Я чат-бот сети салонов <i>Инстинкт</i>\n\n 🌟 Лучший релакс в твоем '
                         'городе\n\n Выберите город', reply_markup=city_selection_key, disable_web_page_preview=True,
                         parse_mode=types.ParseMode.HTML)


@dp.message_handler(commands=['export_bonus'])
async def export_command(message: types.Message):
    """Обработчик команды /export_bonus"""
    # Проверяем, является ли пользователь, который вызывает команду, администратором
    if message.from_user.id not in [5837917794, 5958542955]:  # Предоставление доступа к команде  /export_bonus
        await message.reply('У вас нет доступа к этой команде.')
        return

    data = get_export_bonus_from_database()  # Получаем данные бонусов

    wb = openpyxl.Workbook()  # Создаем файл Excel и записываем данные
    sheet = wb.active
    # Записываем заголовки
    headers = ['user_key', 'id', 'full_name', 'user_name', 'bonus', 'plase']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header
    # Записываем данные пользователей
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            sheet[f'{col_letter}{row_num}'] = cell_data

    wb.save('users_bonus.xlsx')  # Сохраняем файл Excel
    # Отправляем файл пользователю
    with open('users_bonus.xlsx', 'rb') as file:
        await bot.send_document(message.from_user.id, file, caption='Данные пользователей в формате Excel')

    os.remove('users_bonus.xlsx')  # Удаляем файл Excel


@dp.message_handler(commands=['export_user'])
async def export_command(message: types.Message):
    """Обработчик команды /export_user"""
    # Проверяем, является ли пользователь, который вызывает команду, администратором
    if message.from_user.id not in [5837917794, 5958542955]:  # Предоставление доступа к команде  /export_user
        await message.reply('У вас нет доступа к этой команде.')
        return

    data = get_export_user_bonus_from_database()

    wb = openpyxl.Workbook()  # Создаем файл Excel и записываем данные
    sheet = wb.active
    # Записываем заголовки
    headers = ['ID', 'User ID', 'First Name', 'Last Name', 'Username', 'Date']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header
    # Записываем данные пользователей
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            sheet[f'{col_letter}{row_num}'] = cell_data

    wb.save('users.xlsx')  # Сохраняем файл Excel
    # Отправляем файл пользователю
    with open('users.xlsx', 'rb') as file:
        await bot.send_document(message.from_user.id, file, caption='Данные пользователей в формате Excel')

    os.remove('users.xlsx')  # Удаляем файл Excel


@dp.callback_query_handler(lambda c: c.data == "get_a_bonus")
async def get_a_bonus(callback_query: types.CallbackQuery):
    bonus_keyboard = bonus_keyboards()
    bonus_posts = 'Выберете филиал:'
    await bot.send_message(callback_query.from_user.id, bonus_posts, reply_markup=bonus_keyboard,
                           parse_mode=types.ParseMode.HTML)


@dp.callback_query_handler(lambda c: c.data == "top_pard")
async def get_a_bonus(callback_query: types.CallbackQuery):
    top_kub_keyboard = top_kub_keyboards()
    await bot.send_message(callback_query.from_user.id, bonus_post, reply_markup=top_kub_keyboard,
                           parse_mode=types.ParseMode.HTML)


@dp.callback_query_handler(lambda c: c.data == "bottom_part")
async def get_a_bonus(callback_query: types.CallbackQuery):
    bottom_kub_keyboard = bottom_kub_keyboards()
    await bot.send_message(callback_query.from_user.id, bonus_post, reply_markup=bottom_kub_keyboard,
                           parse_mode=types.ParseMode.HTML)


class MakingAnOrder(StatesGroup):
    """Создание класса состояний"""
    write_phone = State()


@dp.callback_query_handler(lambda c: c.data == "top_kub")
async def share_number(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.send_dice(callback_query.from_user.id, emoji='🎲')  # Отправляем эмодзи '🎲'
    time.sleep(5)
    user_id = callback_query.from_user.id
    today = date.today().strftime('%Y-%m-%d')
    plase = "Добролюбова, д.4"
    # Проверка, был ли использован бонус сегодня
    user_key = f"{user_id}_{today}"  # Создание уникального ключа

    existing_user = retrieve_user_bonus(user_key)

    if existing_user:
        text_error_bonus = "Вы уже использовали бонус сегодня."
        await bot.answer_callback_query(callback_query.id, text_error_bonus)
        await state.finish()
        return
    text = "✅ Введите ваше имя."
    await bot.send_message(callback_query.from_user.id, text)
    await MakingAnOrder.write_phone.set()
    await state.update_data(user_id=user_id, today=today, plase=plase)


@dp.callback_query_handler(lambda c: c.data == "bottom_kub")
async def share_number(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.send_dice(callback_query.from_user.id, emoji='🎲')  # Отправляем эмодзи '🎲'
    time.sleep(5)
    user_id = callback_query.from_user.id
    today = date.today().strftime('%Y-%m-%d')
    plase = "проспект Ильича, д.25"
    # Проверка, был ли использован бонус сегодня
    user_key = f"{user_id}_{today}"  # Создание уникального ключа

    existing_user = retrieve_user_bonus(user_key)

    if existing_user:
        text_error_bonus = "Вы уже использовали бонус сегодня."
        await bot.answer_callback_query(callback_query.id, text_error_bonus)
        await state.finish()
        return
    text = "✅ Введите ваше имя."
    await bot.send_message(callback_query.from_user.id, text)
    await MakingAnOrder.write_phone.set()
    await state.update_data(user_id=user_id, today=today, plase=plase)


@dp.message_handler(state=MakingAnOrder.write_phone)
async def write_phone(message: types.Message, state: FSMContext):
    """Обработчик ввода номера телефона"""
    data = await state.get_data()
    user_id = data.get('user_id')
    plase = data.get('plase')
    today = data.get('today')
    phone = message.text
    random_bonus = random.choice(random_bon)
    user_key = f"{user_id}_{today}"  # Создание уникального ключа
    # Проверяем, существует ли запись с таким же id и user_key в таблице users_bonus
    conn = sqlite3.connect('orders.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM users_bonus WHERE id=? AND user_key=?", (user_id, user_key))
    existing_user = cursor.fetchone()
    if existing_user:
        text_error_bonus = "Вы уже использовали бонус сегодня."
        await bot.send_message(message.from_user.id, text_error_bonus)
        await state.finish()
        return
    cursor.execute(
        "INSERT INTO users_bonus (user_key, id, full_name, user_name, bonus, plase) VALUES (?, ?, ?, ?, ?, ?)",
        (user_key, user_id, message.from_user.full_name, phone, random_bonus, plase))

    conn.commit()
    await state.finish()
    bonus = (f"🎉 Ура! А вот и твоя награда: {random_bonus}\n\n"

             f"Предъяви эту запись нашему администратору и забирай свой приз 🏆\n\n"

             f"⚠️ ВНИМАНИЕ! Новый бонус можно получить ровно через сутки.\n"
             f"📢 Подписывайся на наш <a href='https://t.me/instinkt_project_nn'>Telegram канал!</a>\n\n"

             f"Для возврата в начало нажми /start.")
    await message.answer(bonus, disable_web_page_preview=True)


def greeting_handler():
    """Регистрируем handlers для калькулятора"""
    dp.register_message_handler(greeting)  # Обработчик команды /start, он же пост приветствия
